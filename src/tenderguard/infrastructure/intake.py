from __future__ import annotations

import hashlib
import mimetypes
import re
import tempfile
import zipfile
from codecs import getincrementaldecoder
from collections.abc import Callable, Iterable
from io import BytesIO
from pathlib import PurePosixPath
from typing import Any, BinaryIO, cast
from urllib.parse import urlsplit
from xml.parsers import expat

from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from PIL import Image, UnidentifiedImageError
from pypdf import PdfReader

from tenderguard.config import Settings
from tenderguard.domain.common import content_hash
from tenderguard.domain.enums import Severity
from tenderguard.domain.intake import (
    FileInspection,
    IntakeFinding,
    IntakeManifest,
    SheetInspection,
)
from tenderguard.infrastructure.object_store import copy_limited


class _Budget:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        self.files = 0
        self.unpacked_bytes = 0

    def consume(self, size: int) -> None:
        self.files += 1
        self.unpacked_bytes += size
        if self.files > self.settings.max_archive_files:
            raise ValueError("Archive file-count limit exceeded")
        if self.unpacked_bytes > self.settings.max_archive_unpacked_bytes:
            raise ValueError("Archive uncompressed-size limit exceeded")


_ARCHIVE_SUFFIXES = {".zip"}
_UNSUPPORTED_ARCHIVE_SUFFIXES = {".7z", ".rar", ".tar", ".gz", ".bz2", ".xz"}
_OFFICE_CONTAINER_SUFFIXES = {".docx", ".xlsx", ".xlsm", ".pptx"}
_SUPPORTED_SUFFIXES = {
    ".csv",
    ".docx",
    ".dwg",
    ".dxf",
    ".ifc",
    ".jpeg",
    ".jpg",
    ".json",
    ".pdf",
    ".png",
    ".pptx",
    ".tif",
    ".tiff",
    ".txt",
    ".xls",
    ".xlsm",
    ".xlsx",
    ".xml",
    *_ARCHIVE_SUFFIXES,
    *_UNSUPPORTED_ARCHIVE_SUFFIXES,
}
_IMAGE_SUFFIXES = {".jpeg", ".jpg", ".png", ".tif", ".tiff"}
_TEXT_SUFFIXES = {".csv", ".txt"}
_QUALIFIED_ADAPTER_SUFFIXES = {".dwg", ".dxf", ".ifc", ".json", ".xml"}
_MAX_REPORTED_EXCEL_ERROR_CELLS = 50
_MAX_REPORTED_OFFICE_DETAILS = 50
_OFFICE_MAIN_PARTS = {
    ".docx": "word/document.xml",
    ".xlsx": "xl/workbook.xml",
    ".xlsm": "xl/workbook.xml",
    ".pptx": "ppt/presentation.xml",
}
_OFFICE_MAIN_CONTENT_TYPES = {
    ".docx": frozenset(
        {"application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"}
    ),
    ".xlsx": frozenset(
        {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"}
    ),
    ".xlsm": frozenset({"application/vnd.ms-excel.sheet.macroEnabled.main+xml"}),
    ".pptx": frozenset(
        {"application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"}
    ),
}
_OFFICE_MAIN_ROOTS = {
    ".docx": frozenset(
        {
            "http://schemas.openxmlformats.org/wordprocessingml/2006/main}document",
            "http://purl.oclc.org/ooxml/wordprocessingml/main}document",
        }
    ),
    ".xlsx": frozenset(
        {
            "http://schemas.openxmlformats.org/spreadsheetml/2006/main}workbook",
            "http://purl.oclc.org/ooxml/spreadsheetml/main}workbook",
        }
    ),
    ".xlsm": frozenset(
        {
            "http://schemas.openxmlformats.org/spreadsheetml/2006/main}workbook",
            "http://purl.oclc.org/ooxml/spreadsheetml/main}workbook",
        }
    ),
    ".pptx": frozenset(
        {
            "http://schemas.openxmlformats.org/presentationml/2006/main}presentation",
            "http://purl.oclc.org/ooxml/presentationml/main}presentation",
        }
    ),
}
_CONTENT_TYPES_ROOT = "http://schemas.openxmlformats.org/package/2006/content-types}Types"
_CONTENT_TYPE_OVERRIDE = "http://schemas.openxmlformats.org/package/2006/content-types}Override"
_RELATIONSHIPS_ROOTS = frozenset(
    {
        "http://schemas.openxmlformats.org/package/2006/relationships}Relationships",
        "http://purl.oclc.org/ooxml/package/relationships}Relationships",
    }
)
_RELATIONSHIP_NAMES = frozenset(
    {
        "http://schemas.openxmlformats.org/package/2006/relationships}Relationship",
        "http://purl.oclc.org/ooxml/package/relationships}Relationship",
    }
)
_OFFICE_DOCUMENT_RELATIONSHIP_TYPES = frozenset(
    {
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument",
        "http://purl.oclc.org/ooxml/officeDocument/relationships/officeDocument",
    }
)
_HYPERLINK_RELATIONSHIP_TYPES = frozenset(
    {
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink",
        "http://purl.oclc.org/ooxml/officeDocument/relationships/hyperlink",
    }
)
_ORDINARY_EXTERNAL_HYPERLINK_SCHEMES = frozenset({"http", "https", "mailto"})


def _safe_archive_path(name: str) -> bool:
    normalized = name.replace("\\", "/")
    path = PurePosixPath(normalized)
    return (
        not path.is_absolute()
        and ".." not in path.parts
        and not re.match(r"^[A-Za-z]:", normalized)
    )


def _media_type(name: str) -> str:
    return mimetypes.guess_type(name)[0] or "application/octet-stream"


def _finding(
    code: str,
    severity: Severity,
    archive_path: str,
    message: str,
    **details: Any,
) -> IntakeFinding:
    return IntakeFinding(
        code=code,
        severity=severity,
        archive_path=archive_path,
        message=message,
        details=details,
    )


def _rewind(stream: BinaryIO) -> None:
    stream.seek(0)


def _hash_stream(stream: BinaryIO) -> tuple[str, int]:
    _rewind(stream)
    digest = hashlib.sha256()
    size = 0
    while chunk := stream.read(1024 * 1024):
        digest.update(chunk)
        size += len(chunk)
    _rewind(stream)
    return digest.hexdigest(), size


def _starts_with(stream: BinaryIO, signatures: tuple[bytes, ...]) -> bool:
    _rewind(stream)
    prefix = stream.read(max(len(signature) for signature in signatures))
    _rewind(stream)
    return any(prefix.startswith(signature) for signature in signatures)


def _inspect_pdf(path: str, stream: BinaryIO) -> tuple[int | None, int, bool, list[IntakeFinding]]:
    findings: list[IntakeFinding] = []
    if not _starts_with(stream, (b"%PDF-",)):
        return (
            None,
            0,
            False,
            [
                _finding(
                    "FILE_SIGNATURE_MISMATCH",
                    Severity.BLOCKER,
                    path,
                    "PDF extension does not match the file signature",
                )
            ],
        )
    try:
        _rewind(stream)
        reader = PdfReader(stream, strict=True)
        protected = bool(reader.is_encrypted)
        if protected and reader.decrypt("") == 0:
            return (
                None,
                0,
                True,
                [
                    _finding(
                        "PROTECTED_PDF",
                        Severity.BLOCKER,
                        path,
                        "PDF is encrypted and cannot be inspected",
                    )
                ],
            )
        page_count = len(reader.pages)
        attachments = reader.attachments
        embedded_count = len(attachments)
        if embedded_count:
            findings.append(
                _finding(
                    "PDF_EMBEDDED_FILES",
                    Severity.BLOCKER,
                    path,
                    "PDF contains embedded files that require separate manifest entries",
                    count=embedded_count,
                )
            )
        return page_count, embedded_count, protected, findings
    except Exception as error:
        return (
            None,
            0,
            False,
            [
                _finding(
                    "CORRUPT_PDF",
                    Severity.BLOCKER,
                    path,
                    "PDF parser failed",
                    error_type=type(error).__name__,
                )
            ],
        )


def _inspect_excel(
    path: str, stream: BinaryIO
) -> tuple[tuple[SheetInspection, ...], list[IntakeFinding]]:
    findings: list[IntakeFinding] = []
    if not _starts_with(stream, (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")):
        raise ValueError("Excel extension does not match the ZIP container signature")
    _rewind(stream)
    formulas = load_workbook(stream, data_only=False, read_only=False, keep_links=True)
    _rewind(stream)
    cached = load_workbook(stream, data_only=True, read_only=False, keep_links=True)
    sheets: list[SheetInspection] = []
    try:
        for worksheet in formulas.worksheets:
            cached_sheet = cached[worksheet.title]
            formula_count = 0
            missing_cached = 0
            formula_errors: list[dict[str, Any]] = []
            non_formula_errors: list[dict[str, str]] = []
            for row in worksheet.iter_rows():
                for cell in row:
                    cached_cell = cached_sheet[cell.coordinate]
                    if cell.data_type == "f":
                        formula_count += 1
                        if cached_cell.value is None:
                            missing_cached += 1
                        error_signals: set[str] = set()
                        try:
                            formula_value = getattr(cell.value, "text", cell.value)
                            error_signals.update(
                                token.value
                                for token in Tokenizer(str(formula_value)).items
                                if token.type == "OPERAND" and token.subtype == "ERROR"
                            )
                        except Exception as error:
                            error_signals.add(f"UNPARSEABLE:{type(error).__name__}")
                        if cached_cell.data_type == "e":
                            error_signals.add(str(cached_cell.value))
                        if error_signals:
                            formula_errors.append(
                                {
                                    "coordinate": cell.coordinate,
                                    "errors": sorted(error_signals),
                                }
                            )
                    elif cell.data_type == "e":
                        non_formula_errors.append(
                            {
                                "coordinate": cell.coordinate,
                                "error": str(cell.value),
                            }
                        )
            hidden_rows = sum(
                1 for dimension in worksheet.row_dimensions.values() if dimension.hidden
            )
            hidden_column_indexes: set[int] = set()
            for dimension in worksheet.column_dimensions.values():
                if not dimension.hidden:
                    continue
                start = dimension.min
                end = dimension.max
                if start is None or end is None:
                    raise ValueError("Hidden Excel column range lacks bounds")
                hidden_column_indexes.update(range(start, end + 1))
            hidden_columns = len(hidden_column_indexes)
            inspection = SheetInspection(
                name=worksheet.title,
                state=worksheet.sheet_state,
                max_row=worksheet.max_row,
                max_column=worksheet.max_column,
                hidden_row_count=hidden_rows,
                hidden_column_count=hidden_columns,
                formula_cell_count=formula_count,
                formula_without_cached_value_count=missing_cached,
                formula_error_cell_count=len(formula_errors),
                non_formula_error_cell_count=len(non_formula_errors),
            )
            sheets.append(inspection)
            if worksheet.sheet_state != "visible":
                findings.append(
                    _finding(
                        "HIDDEN_EXCEL_SHEET",
                        Severity.WARNING,
                        path,
                        f"Workbook contains {worksheet.sheet_state} sheet",
                        sheet=worksheet.title,
                    )
                )
            if hidden_rows or hidden_columns:
                findings.append(
                    _finding(
                        "HIDDEN_EXCEL_DIMENSIONS",
                        Severity.WARNING,
                        path,
                        "Workbook contains hidden rows or columns",
                        sheet=worksheet.title,
                        hidden_rows=hidden_rows,
                        hidden_columns=hidden_columns,
                    )
                )
            if missing_cached:
                findings.append(
                    _finding(
                        "EXCEL_FORMULA_CACHE_MISSING",
                        Severity.BLOCKER,
                        path,
                        "Formula cells lack cached calculated values",
                        sheet=worksheet.title,
                        count=missing_cached,
                    )
                )
            if formula_errors:
                findings.append(
                    _finding(
                        "EXCEL_FORMULA_ERROR",
                        Severity.BLOCKER,
                        path,
                        "Formula cells contain error tokens or cached error results",
                        sheet=worksheet.title,
                        count=len(formula_errors),
                        cells=formula_errors[:_MAX_REPORTED_EXCEL_ERROR_CELLS],
                        cells_truncated=(len(formula_errors) > _MAX_REPORTED_EXCEL_ERROR_CELLS),
                    )
                )
            if non_formula_errors:
                findings.append(
                    _finding(
                        "EXCEL_CELL_ERROR",
                        Severity.BLOCKER,
                        path,
                        "Workbook contains non-formula error values",
                        sheet=worksheet.title,
                        count=len(non_formula_errors),
                        cells=non_formula_errors[:_MAX_REPORTED_EXCEL_ERROR_CELLS],
                        cells_truncated=(len(non_formula_errors) > _MAX_REPORTED_EXCEL_ERROR_CELLS),
                    )
                )
        external_links = getattr(formulas, "_external_links", [])
        if external_links:
            findings.append(
                _finding(
                    "EXCEL_EXTERNAL_LINKS",
                    Severity.BLOCKER,
                    path,
                    "Workbook contains external links whose values may be stale or unavailable",
                    count=len(external_links),
                )
            )
        return tuple(sheets), findings
    finally:
        formulas.close()
        cached.close()


def _inspect_image(path: str, stream: BinaryIO) -> list[IntakeFinding]:
    try:
        _rewind(stream)
        with Image.open(stream) as image:
            image.verify()
        _rewind(stream)
        return []
    except (UnidentifiedImageError, OSError, Image.DecompressionBombError) as error:
        return [
            _finding(
                "CORRUPT_OR_UNSAFE_IMAGE",
                Severity.BLOCKER,
                path,
                "Image could not be safely decoded",
                error_type=type(error).__name__,
            )
        ]


def _parse_office_xml_part(
    package: zipfile.ZipFile,
    info: zipfile.ZipInfo,
    *,
    on_start: Callable[[str, dict[str, str]], None] | None = None,
) -> str:
    root_name: str | None = None
    parser = expat.ParserCreate(namespace_separator="}")
    parser.buffer_text = True
    parser.SetParamEntityParsing(expat.XML_PARAM_ENTITY_PARSING_NEVER)

    def start_element(name: str, attributes: dict[str, str]) -> None:
        nonlocal root_name
        if root_name is None:
            root_name = name
        if on_start is not None:
            on_start(name, attributes)

    def reject_unsafe_declaration(*_arguments: object) -> None:
        raise ValueError("Office XML contains a forbidden declaration")

    def reject_external_entity(
        _context: str,
        _base: str | None,
        _system_id: str | None,
        _public_id: str | None,
    ) -> int:
        raise ValueError("Office XML contains a forbidden external entity")

    parser.StartElementHandler = start_element
    parser.StartDoctypeDeclHandler = reject_unsafe_declaration
    parser.EntityDeclHandler = reject_unsafe_declaration
    parser.UnparsedEntityDeclHandler = reject_unsafe_declaration
    parser.NotationDeclHandler = reject_unsafe_declaration
    parser.ExternalEntityRefHandler = reject_external_entity
    copied = 0
    with package.open(info, "r") as member:
        while chunk := member.read(64 * 1024):
            copied += len(chunk)
            if copied > info.file_size:
                raise RuntimeError("Office XML member exceeds its declared size")
            parser.Parse(chunk, False)
        parser.Parse(b"", True)
    if copied != info.file_size:
        raise RuntimeError("Office XML member size differs from central directory")
    if root_name is None:
        raise ValueError("Office XML part is empty")
    return root_name


def _normalize_root_relationship_target(target: str) -> str | None:
    normalized = target.replace("\\", "/")
    while normalized.startswith("./"):
        normalized = normalized[2:]
    normalized = normalized.removeprefix("/")
    if (
        not normalized
        or "?" in normalized
        or "#" in normalized
        or not _safe_archive_path(normalized)
    ):
        return None
    return PurePosixPath(normalized).as_posix()


def _inspect_office_structure(
    *,
    path: str,
    package: zipfile.ZipFile,
    members: tuple[zipfile.ZipInfo, ...],
) -> tuple[int, int, list[IntakeFinding]]:
    suffix = PurePosixPath(path.casefold()).suffix
    main_part = _OFFICE_MAIN_PARTS[suffix]
    required_parts = frozenset({"[Content_Types].xml", "_rels/.rels", main_part})
    members_by_name: dict[str, list[zipfile.ZipInfo]] = {}
    case_variants: dict[str, set[str]] = {}
    for info in members:
        members_by_name.setdefault(info.filename, []).append(info)
        case_variants.setdefault(info.filename.casefold(), set()).add(info.filename)

    structure_error_count = 0
    structure_errors: list[dict[str, Any]] = []

    def record_structure_error(code: str, **details: Any) -> None:
        nonlocal structure_error_count
        structure_error_count += 1
        if len(structure_errors) < _MAX_REPORTED_OFFICE_DETAILS:
            structure_errors.append({"code": code, **details})

    duplicate_parts = sorted(name for name, entries in members_by_name.items() if len(entries) > 1)
    if duplicate_parts:
        record_structure_error(
            "DUPLICATE_PARTS",
            count=len(duplicate_parts),
            parts=duplicate_parts[:_MAX_REPORTED_OFFICE_DETAILS],
            parts_truncated=len(duplicate_parts) > _MAX_REPORTED_OFFICE_DETAILS,
        )
    colliding_parts = [sorted(variants) for variants in case_variants.values() if len(variants) > 1]
    if colliding_parts:
        record_structure_error(
            "CASE_COLLIDING_PARTS",
            count=len(colliding_parts),
            parts=colliding_parts[:_MAX_REPORTED_OFFICE_DETAILS],
            parts_truncated=len(colliding_parts) > _MAX_REPORTED_OFFICE_DETAILS,
        )
    missing_parts = sorted(required_parts.difference(members_by_name))
    if missing_parts:
        record_structure_error("MISSING_REQUIRED_PARTS", parts=missing_parts)

    parsed_xml_parts: set[str] = set()
    content_type_overrides: dict[str, list[str]] = {}
    content_types_info = members_by_name.get("[Content_Types].xml", [])
    if len(content_types_info) == 1:
        parsed_xml_parts.add("[Content_Types].xml")

        def content_type_start(name: str, attributes: dict[str, str]) -> None:
            if name != _CONTENT_TYPE_OVERRIDE:
                return
            part_name = attributes.get("PartName", "")
            content_type = attributes.get("ContentType", "")
            content_type_overrides.setdefault(part_name, []).append(content_type)

        try:
            root_name = _parse_office_xml_part(
                package,
                content_types_info[0],
                on_start=content_type_start,
            )
            if root_name != _CONTENT_TYPES_ROOT:
                record_structure_error(
                    "INVALID_CONTENT_TYPES_ROOT",
                    part="[Content_Types].xml",
                )
            expected_values = content_type_overrides.get(f"/{main_part}", [])
            if (
                len(expected_values) != 1
                or expected_values[0] not in _OFFICE_MAIN_CONTENT_TYPES[suffix]
            ):
                record_structure_error(
                    "INVALID_MAIN_CONTENT_TYPE",
                    part=main_part,
                    declared=expected_values,
                )
        except (expat.ExpatError, RuntimeError, ValueError) as error:
            record_structure_error(
                "UNSAFE_OR_INVALID_XML",
                part="[Content_Types].xml",
                error_type=type(error).__name__,
            )

    external_hyperlink_count = 0
    external_dependency_count = 0
    hyperlink_samples: list[dict[str, str]] = []
    dependency_samples: list[dict[str, str]] = []
    root_office_targets: list[str] = []

    def make_relationship_start(
        relationship_info: zipfile.ZipInfo,
        relationship_ids: set[str],
    ) -> Callable[[str, dict[str, str]], None]:
        def relationship_start(name: str, attributes: dict[str, str]) -> None:
            nonlocal external_dependency_count, external_hyperlink_count
            if name not in _RELATIONSHIP_NAMES:
                return
            relationship_id = attributes.get("Id", "")
            relationship_type = attributes.get("Type", "")
            target = attributes.get("Target", "")
            target_mode = attributes.get("TargetMode", "Internal")
            if not relationship_id or not relationship_type or not target:
                record_structure_error(
                    "INCOMPLETE_RELATIONSHIP",
                    part=relationship_info.filename,
                )
                return
            if relationship_id in relationship_ids:
                record_structure_error(
                    "DUPLICATE_RELATIONSHIP_ID",
                    part=relationship_info.filename,
                    relationship_id=relationship_id,
                )
            relationship_ids.add(relationship_id)
            if target_mode.casefold() not in {"internal", "external"}:
                record_structure_error(
                    "INVALID_RELATIONSHIP_MODE",
                    part=relationship_info.filename,
                    relationship_id=relationship_id,
                )
                return
            if (
                relationship_info.filename == "_rels/.rels"
                and relationship_type in _OFFICE_DOCUMENT_RELATIONSHIP_TYPES
                and target_mode.casefold() == "internal"
            ):
                normalized = _normalize_root_relationship_target(target)
                if normalized is None:
                    record_structure_error(
                        "INVALID_MAIN_RELATIONSHIP_TARGET",
                        part=relationship_info.filename,
                        relationship_id=relationship_id,
                    )
                else:
                    root_office_targets.append(normalized)
            if target_mode.casefold() != "external":
                return
            try:
                target_scheme = urlsplit(target).scheme.casefold() or "relative"
            except ValueError:
                target_scheme = "invalid"
                record_structure_error(
                    "INVALID_EXTERNAL_TARGET",
                    part=relationship_info.filename,
                    relationship_id=relationship_id,
                )
            relationship_record = {
                "relationship_part": relationship_info.filename,
                "relationship_type": relationship_type.rsplit("/", 1)[-1],
                "target_scheme": target_scheme,
                "target_sha256": hashlib.sha256(target.encode("utf-8")).hexdigest(),
            }
            if (
                relationship_type in _HYPERLINK_RELATIONSHIP_TYPES
                and target_scheme in _ORDINARY_EXTERNAL_HYPERLINK_SCHEMES
            ):
                external_hyperlink_count += 1
                if len(hyperlink_samples) < _MAX_REPORTED_OFFICE_DETAILS:
                    hyperlink_samples.append(relationship_record)
            else:
                external_dependency_count += 1
                if len(dependency_samples) < _MAX_REPORTED_OFFICE_DETAILS:
                    dependency_samples.append(relationship_record)

        return relationship_start

    relationship_infos = tuple(
        info for info in members if info.filename.casefold().endswith(".rels")
    )
    for info in relationship_infos:
        relationship_ids: set[str] = set()

        try:
            root_name = _parse_office_xml_part(
                package,
                info,
                on_start=make_relationship_start(info, relationship_ids),
            )
            if root_name not in _RELATIONSHIPS_ROOTS:
                record_structure_error(
                    "INVALID_RELATIONSHIPS_ROOT",
                    part=info.filename,
                )
        except (expat.ExpatError, RuntimeError, ValueError) as error:
            record_structure_error(
                "UNSAFE_OR_INVALID_XML",
                part=info.filename,
                error_type=type(error).__name__,
            )

    if root_office_targets != [main_part]:
        record_structure_error(
            "INVALID_MAIN_RELATIONSHIP",
            expected=main_part,
            declared=root_office_targets,
        )

    main_part_info = members_by_name.get(main_part, [])
    if len(main_part_info) == 1:
        parsed_xml_parts.add(main_part)
        try:
            root_name = _parse_office_xml_part(package, main_part_info[0])
            if root_name not in _OFFICE_MAIN_ROOTS[suffix]:
                record_structure_error(
                    "INVALID_MAIN_PART_ROOT",
                    part=main_part,
                )
        except (expat.ExpatError, RuntimeError, ValueError) as error:
            record_structure_error(
                "UNSAFE_OR_INVALID_XML",
                part=main_part,
                error_type=type(error).__name__,
            )

    for info in members:
        if not info.filename.casefold().endswith(".xml") or info.filename in parsed_xml_parts:
            continue
        parsed_xml_parts.add(info.filename)
        try:
            _parse_office_xml_part(package, info)
        except (expat.ExpatError, RuntimeError, ValueError) as error:
            record_structure_error(
                "UNSAFE_OR_INVALID_XML",
                part=info.filename,
                error_type=type(error).__name__,
            )

    findings: list[IntakeFinding] = []
    if structure_error_count:
        findings.append(
            _finding(
                "CORRUPT_OFFICE_STRUCTURE",
                Severity.BLOCKER,
                path,
                "Office package structure does not safely match its file type",
                suffix=suffix,
                error_count=structure_error_count,
                errors=structure_errors,
                errors_truncated=structure_error_count > len(structure_errors),
            )
        )
    if external_dependency_count:
        findings.append(
            _finding(
                "OFFICE_EXTERNAL_DEPENDENCY",
                Severity.BLOCKER,
                path,
                "Office package contains non-hyperlink external relationships",
                count=external_dependency_count,
                relationships=dependency_samples,
                relationships_truncated=(external_dependency_count > len(dependency_samples)),
            )
        )
    if external_hyperlink_count:
        findings.append(
            _finding(
                "OFFICE_EXTERNAL_HYPERLINK",
                Severity.WARNING,
                path,
                "Office package contains external hyperlinks",
                count=external_hyperlink_count,
                relationships=hyperlink_samples,
                relationships_truncated=(external_hyperlink_count > len(hyperlink_samples)),
            )
        )
    return external_hyperlink_count, external_dependency_count, findings


def _inspect_office_container(
    path: str,
    stream: BinaryIO,
    settings: Settings,
) -> tuple[int, int, int, list[IntakeFinding]]:
    findings: list[IntakeFinding] = []
    if not _starts_with(stream, (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")):
        return (
            0,
            0,
            0,
            [
                _finding(
                    "FILE_SIGNATURE_MISMATCH",
                    Severity.BLOCKER,
                    path,
                    "Office extension does not match the ZIP container signature",
                )
            ],
        )
    try:
        _rewind(stream)
        with zipfile.ZipFile(stream) as package:
            members = tuple(info for info in package.infolist() if not info.is_dir())
            if len(members) > settings.max_archive_files:
                findings.append(
                    _finding(
                        "OFFICE_FILE_COUNT_EXCEEDED",
                        Severity.BLOCKER,
                        path,
                        "Office package exceeds the configured member-count limit",
                        count=len(members),
                    )
                )
            total_unpacked = sum(info.file_size for info in members)
            if total_unpacked > settings.max_archive_unpacked_bytes:
                findings.append(
                    _finding(
                        "OFFICE_UNPACKED_SIZE_EXCEEDED",
                        Severity.BLOCKER,
                        path,
                        "Office package exceeds the configured unpacked-size limit",
                        size_bytes=total_unpacked,
                    )
                )
            for info in members:
                member_path = f"{path}!/{info.filename}"
                if not _safe_archive_path(info.filename):
                    findings.append(
                        _finding(
                            "OFFICE_PATH_TRAVERSAL",
                            Severity.BLOCKER,
                            member_path,
                            "Office package member path is unsafe",
                        )
                    )
                ratio = (
                    settings.max_archive_compression_ratio + 1
                    if info.compress_size == 0 and info.file_size > 0
                    else info.file_size / max(info.compress_size, 1)
                )
                if ratio > settings.max_archive_compression_ratio:
                    findings.append(
                        _finding(
                            "OFFICE_COMPRESSION_RATIO_EXCEEDED",
                            Severity.BLOCKER,
                            member_path,
                            "Office member exceeds the compression-ratio safety limit",
                            ratio=str(ratio),
                        )
                    )
                if info.flag_bits & 0x1:
                    findings.append(
                        _finding(
                            "PROTECTED_OFFICE_MEMBER",
                            Severity.BLOCKER,
                            member_path,
                            "Office package member is encrypted",
                        )
                    )
            embedded = tuple(info.filename for info in members if "/embeddings/" in info.filename)
            if embedded:
                findings.append(
                    _finding(
                        "OFFICE_EMBEDDED_OBJECTS",
                        Severity.BLOCKER,
                        path,
                        "Office file contains embedded objects requiring separate inspection",
                        count=len(embedded),
                    )
                )
            macro_members = tuple(
                info.filename
                for info in members
                if info.filename.casefold().endswith("vbaproject.bin")
            )
            if macro_members:
                findings.append(
                    _finding(
                        "OFFICE_MACROS_PRESENT",
                        Severity.BLOCKER,
                        path,
                        "Office package contains macro code requiring a qualified review",
                        count=len(macro_members),
                    )
                )
            unsafe_preparse_codes = {
                "OFFICE_FILE_COUNT_EXCEEDED",
                "OFFICE_UNPACKED_SIZE_EXCEEDED",
                "OFFICE_PATH_TRAVERSAL",
                "OFFICE_COMPRESSION_RATIO_EXCEEDED",
                "PROTECTED_OFFICE_MEMBER",
            }
            external_hyperlink_count = 0
            external_dependency_count = 0
            if not any(finding.code in unsafe_preparse_codes for finding in findings):
                (
                    external_hyperlink_count,
                    external_dependency_count,
                    structure_findings,
                ) = _inspect_office_structure(
                    path=path,
                    package=package,
                    members=members,
                )
                findings.extend(structure_findings)
                for info in members:
                    copied = 0
                    with package.open(info, "r") as member:
                        while chunk := member.read(1024 * 1024):
                            copied += len(chunk)
                            if copied > info.file_size:
                                raise RuntimeError("Office member exceeds its declared size")
                    if copied != info.file_size:
                        raise RuntimeError("Office member size differs from central directory")
            return (
                len(embedded),
                external_hyperlink_count,
                external_dependency_count,
                findings,
            )
    except (RuntimeError, zipfile.BadZipFile) as error:
        return (
            0,
            0,
            0,
            [
                _finding(
                    "CORRUPT_OR_PROTECTED_OFFICE_FILE",
                    Severity.BLOCKER,
                    path,
                    "Office package is corrupt, legacy binary, or encrypted",
                    error_type=type(error).__name__,
                )
            ],
        )


def _inspect_text(path: str, stream: BinaryIO) -> list[IntakeFinding]:
    decoder = getincrementaldecoder("utf-8-sig")("strict")
    try:
        _rewind(stream)
        while chunk := stream.read(1024 * 1024):
            if b"\x00" in chunk:
                raise UnicodeError("NUL byte found")
            decoder.decode(chunk)
        decoder.decode(b"", final=True)
        return []
    except UnicodeError:
        return [
            _finding(
                "TEXT_CONTENT_INVALID",
                Severity.BLOCKER,
                path,
                "Text document is not valid UTF-8 text or contains binary NUL bytes",
            )
        ]
    finally:
        _rewind(stream)


def _inspect_single_stream(
    path: str,
    stream: BinaryIO,
    settings: Settings,
    *,
    nested_archive: bool = False,
) -> FileInspection:
    suffix = PurePosixPath(path.casefold()).suffix
    digest, size = _hash_stream(stream)
    findings: list[IntakeFinding] = []
    page_count: int | None = None
    embedded_count = 0
    external_hyperlink_count = 0
    external_dependency_count = 0
    protected = False
    corrupt = False
    unsupported = False
    sheets: tuple[SheetInspection, ...] = ()

    if suffix == ".pdf":
        page_count, embedded_count, protected, pdf_findings = _inspect_pdf(path, stream)
        findings.extend(pdf_findings)
        corrupt = any(item.code == "CORRUPT_PDF" for item in findings)
    elif suffix in {".xlsx", ".xlsm"}:
        (
            embedded_count,
            external_hyperlink_count,
            external_dependency_count,
            office_findings,
        ) = _inspect_office_container(
            path,
            stream,
            settings,
        )
        findings.extend(office_findings)
        corrupt = any(item.code.startswith("CORRUPT_") for item in office_findings)
        protected = any(item.code == "PROTECTED_OFFICE_MEMBER" for item in office_findings)
        if not any(item.severity is Severity.BLOCKER for item in office_findings):
            try:
                sheets, excel_findings = _inspect_excel(path, stream)
                findings.extend(excel_findings)
            except Exception as error:
                corrupt = True
                findings.append(
                    _finding(
                        "CORRUPT_OR_PROTECTED_EXCEL",
                        Severity.BLOCKER,
                        path,
                        "Excel workbook could not be opened",
                        error_type=type(error).__name__,
                    )
                )
    elif suffix in {".docx", ".pptx"}:
        (
            embedded_count,
            external_hyperlink_count,
            external_dependency_count,
            office_findings,
        ) = _inspect_office_container(
            path,
            stream,
            settings,
        )
        findings.extend(office_findings)
        corrupt = any(item.code.startswith("CORRUPT_") for item in office_findings)
        protected = any(item.code == "PROTECTED_OFFICE_MEMBER" for item in office_findings)
    elif suffix in _IMAGE_SUFFIXES:
        image_findings = _inspect_image(path, stream)
        findings.extend(image_findings)
        corrupt = bool(image_findings)
    elif suffix in _TEXT_SUFFIXES:
        text_findings = _inspect_text(path, stream)
        findings.extend(text_findings)
        corrupt = bool(text_findings)
    elif suffix == ".xls":
        unsupported = True
        findings.append(
            _finding(
                "LEGACY_EXCEL_ADAPTER_REQUIRED",
                Severity.BLOCKER,
                path,
                "Legacy XLS requires a qualified conversion/parser adapter",
            )
        )
    elif suffix in _QUALIFIED_ADAPTER_SUFFIXES:
        unsupported = True
        findings.append(
            _finding(
                "DOCUMENT_ADAPTER_REQUIRED",
                Severity.BLOCKER,
                path,
                f"File format {suffix} requires a qualified isolated parser adapter",
            )
        )
    elif suffix in _UNSUPPORTED_ARCHIVE_SUFFIXES:
        unsupported = True
        findings.append(
            _finding(
                "ARCHIVE_ADAPTER_REQUIRED",
                Severity.BLOCKER,
                path,
                f"Archive format {suffix} requires a qualified adapter",
            )
        )
    elif suffix not in _SUPPORTED_SUFFIXES:
        unsupported = True
        findings.append(
            _finding(
                "UNSUPPORTED_FILE_TYPE",
                Severity.BLOCKER,
                path,
                "File type is not in the document-intake allowlist",
                suffix=suffix,
            )
        )

    return FileInspection(
        entry_id=f"file-{content_hash({'path': path, 'sha256': digest})[:24]}",
        archive_path=path,
        sha256=digest,
        size_bytes=size,
        media_type=_media_type(path),
        nested_archive=nested_archive,
        corrupt=corrupt,
        protected=protected,
        unsupported=unsupported,
        page_count=page_count,
        embedded_file_count=embedded_count,
        external_hyperlink_count=external_hyperlink_count,
        external_dependency_count=external_dependency_count,
        sheets=sheets,
        findings=tuple(findings),
    )


def _walk_zip_stream(
    *,
    archive_path: str,
    stream: BinaryIO,
    depth: int,
    budget: _Budget,
    on_member: Callable[[str, BinaryIO], None] | None = None,
) -> tuple[list[FileInspection], list[IntakeFinding]]:
    entries: list[FileInspection] = []
    findings: list[IntakeFinding] = []
    if depth > budget.settings.max_archive_depth:
        findings.append(
            _finding(
                "ARCHIVE_DEPTH_EXCEEDED",
                Severity.BLOCKER,
                archive_path,
                "Nested archive depth exceeds configured safety limit",
            )
        )
        return entries, findings
    try:
        _rewind(stream)
        with zipfile.ZipFile(stream) as archive:
            for info in archive.infolist():
                if info.is_dir():
                    continue
                child_path = f"{archive_path}!/{info.filename}"
                if not _safe_archive_path(info.filename):
                    findings.append(
                        _finding(
                            "ARCHIVE_PATH_TRAVERSAL",
                            Severity.BLOCKER,
                            child_path,
                            "Archive member path is unsafe",
                        )
                    )
                    continue
                budget.consume(info.file_size)
                if info.compress_size == 0 and info.file_size > 0:
                    ratio = float(budget.settings.max_archive_compression_ratio + 1)
                else:
                    ratio = info.file_size / max(info.compress_size, 1)
                if ratio > budget.settings.max_archive_compression_ratio:
                    findings.append(
                        _finding(
                            "ARCHIVE_COMPRESSION_RATIO_EXCEEDED",
                            Severity.BLOCKER,
                            child_path,
                            "Archive member exceeds compression-ratio safety limit",
                            ratio=str(ratio),
                        )
                    )
                    continue
                if info.flag_bits & 0x1:
                    findings.append(
                        _finding(
                            "PROTECTED_ARCHIVE_MEMBER",
                            Severity.BLOCKER,
                            child_path,
                            "Archive member is encrypted",
                        )
                    )
                    entries.append(
                        FileInspection(
                            entry_id=f"file-{content_hash(child_path)[:24]}",
                            archive_path=child_path,
                            sha256="0" * 64,
                            size_bytes=info.file_size,
                            media_type=_media_type(info.filename),
                            protected=True,
                        )
                    )
                    continue
                with (
                    archive.open(info, "r") as member_stream,
                    tempfile.SpooledTemporaryFile(
                        max_size=budget.settings.max_parser_spool_memory_bytes
                    ) as child,
                ):
                    member_binary = cast(BinaryIO, member_stream)
                    child_binary = cast(BinaryIO, child)
                    copied = copy_limited(
                        member_binary,
                        child_binary,
                        min(info.file_size, budget.settings.max_archive_unpacked_bytes),
                    )
                    if copied != info.file_size:
                        raise RuntimeError("Archive member size differs from central directory")
                    _rewind(child_binary)
                    if on_member is not None:
                        on_member(child_path, child_binary)
                        _rewind(child_binary)
                    suffix = PurePosixPath(info.filename.casefold()).suffix
                    is_archive = (
                        suffix in _ARCHIVE_SUFFIXES and suffix not in _OFFICE_CONTAINER_SUFFIXES
                    )
                    inspected = _inspect_single_stream(
                        child_path,
                        child_binary,
                        budget.settings,
                        nested_archive=is_archive,
                    )
                    entries.append(inspected)
                    findings.extend(inspected.findings)
                    if is_archive:
                        nested_entries, nested_findings = _walk_zip_stream(
                            archive_path=child_path,
                            stream=child_binary,
                            depth=depth + 1,
                            budget=budget,
                            on_member=on_member,
                        )
                        entries.extend(nested_entries)
                        findings.extend(nested_findings)
    except (zipfile.BadZipFile, RuntimeError) as error:
        findings.append(
            _finding(
                "CORRUPT_ARCHIVE",
                Severity.BLOCKER,
                archive_path,
                "Archive could not be completely enumerated",
                error_type=type(error).__name__,
            )
        )
    except ValueError as error:
        findings.append(
            _finding(
                "ARCHIVE_LIMIT_EXCEEDED",
                Severity.BLOCKER,
                archive_path,
                str(error),
            )
        )
    return entries, findings


def inspect_intake_stream(
    filename: str,
    stream: BinaryIO,
    settings: Settings,
    *,
    on_member: Callable[[str, BinaryIO], None] | None = None,
) -> IntakeManifest:
    root_hash, root_size = _hash_stream(stream)
    if root_size > settings.max_upload_bytes:
        raise ValueError(f"Upload exceeds configured limit of {settings.max_upload_bytes} bytes")
    budget = _Budget(settings)
    budget.consume(root_size)
    suffix = PurePosixPath(filename.casefold()).suffix
    if suffix in _ARCHIVE_SUFFIXES:
        entries, findings = _walk_zip_stream(
            archive_path=filename,
            stream=stream,
            depth=0,
            budget=budget,
            on_member=on_member,
        )
    else:
        entry = _inspect_single_stream(filename, stream, settings)
        entries = [entry]
        findings = list(entry.findings)
    all_processed = (
        bool(entries)
        and not any(item.corrupt or item.protected or item.unsupported for item in entries)
        and not any(item.severity is Severity.BLOCKER for item in findings)
    )
    return IntakeManifest(
        root_filename=filename,
        root_sha256=root_hash,
        entries=tuple(entries),
        findings=tuple(findings),
        all_files_processed=all_processed,
    )


def inspect_intake(
    filename: str,
    content: bytes,
    settings: Settings,
    *,
    on_member: Callable[[str, bytes], None] | None = None,
) -> IntakeManifest:
    member_callback: Callable[[str, BinaryIO], None] | None = None
    if on_member is not None:

        def member_callback(path: str, stream: BinaryIO) -> None:
            _rewind(stream)
            on_member(path, stream.read())
            _rewind(stream)

    return inspect_intake_stream(
        filename,
        BytesIO(content),
        settings,
        on_member=member_callback,
    )


def missing_referenced_documents(
    *,
    extracted_references: Iterable[str],
    available_logical_keys: Iterable[str],
) -> tuple[str, ...]:
    available = {" ".join(item.casefold().split()) for item in available_logical_keys}
    return tuple(
        sorted(
            reference
            for reference in extracted_references
            if " ".join(reference.casefold().split()) not in available
        )
    )
