from __future__ import annotations

import hashlib
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from tenderguard.domain.boq_spreadsheet import (
    BoqCellEvidence,
    BoqRowCandidate,
    BoqXlsxColumn,
    BoqXlsxExtractionResult,
    BoqXlsxProfile,
)
from tenderguard.domain.common import content_hash, utc_now
from tenderguard.domain.enums import Severity
from tenderguard.domain.intake import FileInspection, IntakeManifest


class BoqXlsxExtractionError(RuntimeError):
    def __init__(self, *, code: str) -> None:
        self.code = code
        super().__init__(code)


def extract_boq_xlsx_candidates(
    *,
    workbook_content: bytes,
    workbook_archive_path: str,
    manifest: IntakeManifest,
    profile: BoqXlsxProfile,
    extracted_at: datetime | None = None,
) -> BoqXlsxExtractionResult:
    workbook_hash = hashlib.sha256(workbook_content).hexdigest()
    if profile.expected_workbook_sha256 not in {None, workbook_hash}:
        raise BoqXlsxExtractionError(code="BOQ_WORKBOOK_FINGERPRINT_MISMATCH")
    entries = tuple(
        entry
        for entry in manifest.entries
        if entry.archive_path == workbook_archive_path and entry.sha256 == workbook_hash
    )
    if len(entries) != 1:
        raise BoqXlsxExtractionError(code="BOQ_MANIFEST_ENTRY_AMBIGUOUS")
    entry = entries[0]
    if entry.corrupt or entry.protected or entry.unsupported:
        raise BoqXlsxExtractionError(code="BOQ_WORKBOOK_UNPROCESSABLE")

    global_blockers = _manifest_blockers(manifest, entry)
    try:
        formulas = load_workbook(
            BytesIO(workbook_content),
            data_only=False,
            read_only=False,
            keep_links=True,
        )
        cached = load_workbook(
            BytesIO(workbook_content),
            data_only=True,
            read_only=False,
            keep_links=True,
        )
    except Exception as error:
        raise BoqXlsxExtractionError(code="BOQ_WORKBOOK_PARSER_FAILED") from error

    try:
        if getattr(formulas, "_external_links", []):
            global_blockers.add("EXCEL_EXTERNAL_LINKS")
        if profile.worksheet_name not in formulas.sheetnames:
            raise BoqXlsxExtractionError(code="BOQ_WORKSHEET_NOT_FOUND")
        if profile.worksheet_name not in cached.sheetnames:
            raise BoqXlsxExtractionError(code="BOQ_CACHED_WORKSHEET_NOT_FOUND")
        worksheet = formulas[profile.worksheet_name]
        cached_worksheet = cached[profile.worksheet_name]
        if worksheet.sheet_state != "visible":
            global_blockers.add("SELECTED_WORKSHEET_NOT_VISIBLE")
        _validate_headers(worksheet, profile)
        if profile.data_end_row > worksheet.max_row:
            raise BoqXlsxExtractionError(code="BOQ_PROFILE_DATA_RANGE_EXCEEDS_SHEET")

        candidates: list[BoqRowCandidate] = []
        for row_number in range(profile.data_start_row, profile.data_end_row + 1):
            candidate = _extract_row(
                worksheet=worksheet,
                cached_worksheet=cached_worksheet,
                row_number=row_number,
                profile=profile,
            )
            if candidate is not None:
                candidates.append(candidate)
        if not candidates:
            global_blockers.add("BOQ_NO_POSITION_CANDIDATES")
        candidates = _mark_duplicate_identifiers(candidates)
        status = (
            "BLOCKED"
            if global_blockers or any(candidate.blockers for candidate in candidates)
            else "UNVERIFIED"
        )
        return BoqXlsxExtractionResult(
            status=status,
            profile_version_id=profile.profile_version_id,
            profile_content_hash=content_hash(profile),
            root_object_sha256=manifest.root_sha256,
            workbook_object_sha256=workbook_hash,
            archive_path=workbook_archive_path,
            worksheet_name=profile.worksheet_name,
            candidates=tuple(candidates),
            global_blockers=tuple(sorted(global_blockers)),
            extracted_at=extracted_at or utc_now(),
        )
    finally:
        formulas.close()
        cached.close()


def _manifest_blockers(
    manifest: IntakeManifest,
    entry: FileInspection,
) -> set[str]:
    blockers: set[str] = set()
    if not manifest.all_files_processed:
        blockers.add("INTAKE_MANIFEST_BLOCKED")
    for finding in (*manifest.findings, *entry.findings):
        if finding.severity is Severity.BLOCKER:
            blockers.add(f"INTAKE_{finding.code}")
    if any(
        sheet.state != "visible" or sheet.hidden_row_count or sheet.hidden_column_count
        for sheet in entry.sheets
    ):
        blockers.add("HIDDEN_WORKBOOK_CONTENT")
    return blockers


def _profile_columns(profile: BoqXlsxProfile) -> dict[str, BoqXlsxColumn]:
    columns = {
        "position_id": profile.position_id,
        "description": profile.description,
        "unit": profile.unit,
        "quantity": profile.quantity,
    }
    if profile.specification is not None:
        columns["specification"] = profile.specification
    if profile.reference is not None:
        columns["reference"] = profile.reference
    return columns


def _validate_headers(
    worksheet: Worksheet,
    profile: BoqXlsxProfile,
) -> None:
    for semantic_name, column in _profile_columns(profile).items():
        cell = worksheet.cell(profile.header_row, column.column)
        if cell.data_type == "f" or cell.value != column.header:
            raise BoqXlsxExtractionError(code=f"BOQ_HEADER_MISMATCH_{semantic_name.upper()}")
        matching_columns = tuple(
            candidate.column
            for candidate in worksheet[profile.header_row]
            if candidate.value == column.header
        )
        if matching_columns != (column.column,):
            raise BoqXlsxExtractionError(code=f"BOQ_HEADER_AMBIGUOUS_{semantic_name.upper()}")
        if _cell_is_merged(worksheet, cell.coordinate):
            raise BoqXlsxExtractionError(code=f"BOQ_HEADER_MERGED_{semantic_name.upper()}")


def _extract_row(
    *,
    worksheet: Worksheet,
    cached_worksheet: Worksheet,
    row_number: int,
    profile: BoqXlsxProfile,
) -> BoqRowCandidate | None:
    columns = _profile_columns(profile)
    formula_cells = {
        name: worksheet.cell(row_number, column.column) for name, column in columns.items()
    }
    cached_cells = {
        name: cached_worksheet.cell(row_number, column.column) for name, column in columns.items()
    }
    if all(cell.value in (None, "") for cell in formula_cells.values()):
        return None

    row_text = " ".join(
        str(cell.value).strip() for cell in formula_cells.values() if cell.value not in (None, "")
    )
    if any(re.search(pattern, row_text) for pattern in profile.section_row_patterns):
        return None

    blockers: set[str] = set()
    evidence = {
        name: _cell_evidence(cell, cached_cells[name]) for name, cell in formula_cells.items()
    }
    for name, cell in formula_cells.items():
        if _cell_is_merged(worksheet, cell.coordinate):
            blockers.add(f"MERGED_{name.upper()}_CELL")

    position_id = _literal_text(
        formula_cells["position_id"],
        field="POSITION_ID",
        blockers=blockers,
    )
    if position_id is None:
        blockers.add("MISSING_STABLE_POSITION_ID")
    elif re.fullmatch(profile.position_id_pattern, position_id) is None:
        blockers.add("POSITION_ID_PATTERN_MISMATCH")

    description = _literal_text(
        formula_cells["description"],
        field="DESCRIPTION",
        blockers=blockers,
    )
    if description is None:
        blockers.add("DESCRIPTION_MISSING")

    unit = _literal_text(
        formula_cells["unit"],
        field="UNIT",
        blockers=blockers,
    )
    if unit is None:
        blockers.add("UNIT_MISSING")
    elif unit not in profile.allowed_units:
        blockers.add("UNIT_NOT_ALLOWED_BY_PROFILE")

    quantity = _quantity(
        formula_cell=formula_cells["quantity"],
        cached_cell=cached_cells["quantity"],
        profile=profile,
        blockers=blockers,
    )

    specification = (
        _literal_text(
            formula_cells["specification"],
            field="SPECIFICATION",
            blockers=blockers,
            required=False,
        )
        if "specification" in formula_cells
        else None
    )
    source_reference = (
        _literal_text(
            formula_cells["reference"],
            field="REFERENCE",
            blockers=blockers,
            required=False,
        )
        if "reference" in formula_cells
        else None
    )
    provisional_identity = {
        "worksheet_name": worksheet.title,
        "row_number": row_number,
        "cells": evidence,
    }
    return BoqRowCandidate(
        provisional_candidate_id=(f"boq-candidate-{content_hash(provisional_identity)[:24]}"),
        worksheet_name=worksheet.title,
        row_number=row_number,
        source_position_id=position_id,
        description=description,
        specification=specification,
        source_reference=source_reference,
        unit=unit,
        quantity=quantity,
        cells=evidence,
        blockers=tuple(sorted(blockers)),
    )


def _literal_text(
    cell: Cell | MergedCell,
    *,
    field: str,
    blockers: set[str],
    required: bool = True,
) -> str | None:
    if cell.data_type == "f":
        blockers.add(f"{field}_FORMULA_NOT_ALLOWED")
        return None
    if cell.value in (None, ""):
        if required:
            blockers.add(f"{field}_MISSING")
        return None
    if not isinstance(cell.value, str | int):
        blockers.add(f"{field}_TYPE_INVALID")
        return None
    value = str(cell.value)
    if value != value.strip() or "\x00" in value:
        blockers.add(f"{field}_LITERAL_INVALID")
        return None
    return value


def _quantity(
    *,
    formula_cell: Cell | MergedCell,
    cached_cell: Cell | MergedCell,
    profile: BoqXlsxProfile,
    blockers: set[str],
) -> Decimal | None:
    value: Any
    if formula_cell.data_type == "f":
        if not profile.allow_quantity_formulas:
            blockers.add("QUANTITY_FORMULA_NOT_ALLOWED")
        value = cached_cell.value
        if value in (None, ""):
            blockers.add("QUANTITY_FORMULA_CACHE_MISSING")
            return None
    else:
        value = formula_cell.value
    if value in (None, ""):
        blockers.add("QUANTITY_MISSING")
        return None
    if isinstance(value, bool):
        blockers.add("QUANTITY_TYPE_INVALID")
        return None
    if isinstance(value, str):
        if value != value.strip():
            blockers.add("QUANTITY_LITERAL_INVALID")
            return None
        literal = value
        if profile.quantity_decimal_separator == ",":
            if "." in literal or literal.count(",") > 1:
                blockers.add("QUANTITY_LITERAL_INVALID")
                return None
            literal = literal.replace(",", ".")
        elif "," in literal:
            blockers.add("QUANTITY_LITERAL_INVALID")
            return None
    elif isinstance(value, int | float | Decimal):
        literal = str(value)
    else:
        blockers.add("QUANTITY_TYPE_INVALID")
        return None
    try:
        quantity = Decimal(literal)
    except InvalidOperation:
        blockers.add("QUANTITY_DECIMAL_INVALID")
        return None
    if not quantity.is_finite() or quantity <= 0:
        blockers.add("QUANTITY_NOT_POSITIVE")
        return None
    exponent = quantity.as_tuple().exponent
    if not isinstance(exponent, int):
        blockers.add("QUANTITY_DECIMAL_INVALID")
        return None
    if len(quantity.as_tuple().digits) > 38 or max(-exponent, 0) > 12:
        blockers.add("QUANTITY_PRECISION_EXCEEDED")
        return None
    return quantity


def _cell_evidence(
    formula_cell: Cell | MergedCell,
    cached_cell: Cell | MergedCell,
) -> BoqCellEvidence:
    value = formula_cell.value
    if formula_cell.data_type == "f":
        return BoqCellEvidence(
            coordinate=formula_cell.coordinate,
            value_kind="FORMULA",
            formula=str(value),
            cached_literal=(None if cached_cell.value is None else str(cached_cell.value)),
        )
    if value in (None, ""):
        kind = "EMPTY"
        literal = None
    elif isinstance(value, bool):
        kind = "BOOLEAN"
        literal = str(value)
    elif isinstance(value, str):
        kind = "TEXT"
        literal = value
    elif isinstance(value, int | float | Decimal):
        kind = "NUMBER"
        literal = str(value)
    else:
        kind = "OTHER"
        literal = str(value)
    return BoqCellEvidence(
        coordinate=formula_cell.coordinate,
        value_kind=kind,
        source_literal=literal,
    )


def _cell_is_merged(worksheet: Worksheet, coordinate: str) -> bool:
    return any(coordinate in merged_range for merged_range in worksheet.merged_cells.ranges)


def _mark_duplicate_identifiers(
    candidates: list[BoqRowCandidate],
) -> list[BoqRowCandidate]:
    counts: dict[str, int] = {}
    for candidate in candidates:
        if candidate.source_position_id is not None:
            counts[candidate.source_position_id] = counts.get(candidate.source_position_id, 0) + 1
    duplicates = {source_position_id for source_position_id, count in counts.items() if count > 1}
    if not duplicates:
        return candidates
    return [
        (
            candidate.model_copy(
                update={"blockers": tuple(sorted({*candidate.blockers, "DUPLICATE_POSITION_ID"}))}
            )
            if candidate.source_position_id in duplicates
            else candidate
        )
        for candidate in candidates
    ]
