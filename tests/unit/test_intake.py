from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

from tenderguard.config import Settings
from tenderguard.domain.intake import IntakeManifest
from tenderguard.infrastructure.intake import inspect_intake, inspect_intake_stream
from tenderguard.infrastructure.object_store import LocalObjectStore


def _workbook_with_cached_formula_error(*, formula: str, cached_error: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cost"
    sheet["A1"] = 10
    sheet["A2"] = formula
    output = BytesIO()
    workbook.save(output)

    formula_xml = formula.removeprefix("=")
    original = f'<c r="A2"><f>{formula_xml}</f><v /></c>'.encode()
    replacement = (f'<c r="A2" t="e"><f>{formula_xml}</f><v>{cached_error}</v></c>').encode()
    patched = BytesIO()
    output.seek(0)
    with (
        ZipFile(output, "r") as source,
        ZipFile(patched, "w", compression=ZIP_DEFLATED) as target,
    ):
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                assert original in payload
                payload = payload.replace(original, replacement)
            target.writestr(info, payload)
    return patched.getvalue()


def _minimal_docx(
    *,
    external_relationship_type: str | None = None,
    external_target: str = "https://example.invalid/reference",
    document_xml: bytes | None = None,
    additional_parts: dict[str, bytes] | None = None,
) -> bytes:
    payload = BytesIO()
    with ZipFile(payload, "w", compression=ZIP_DEFLATED) as package:
        package.writestr(
            "[Content_Types].xml",
            (
                '<?xml version="1.0" encoding="UTF-8"?>'
                '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                '<Override PartName="/word/document.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.'
                'wordprocessingml.document.main+xml"/>'
                "</Types>"
            ),
        )
        package.writestr(
            "_rels/.rels",
            (
                '<?xml version="1.0" encoding="UTF-8"?>'
                "<Relationships "
                'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
                'relationships/officeDocument" Target="word/document.xml"/>'
                "</Relationships>"
            ),
        )
        package.writestr(
            "word/document.xml",
            document_xml
            or (
                b'<?xml version="1.0" encoding="UTF-8"?>'
                b'<w:document xmlns:w="http://schemas.openxmlformats.org/'
                b'wordprocessingml/2006/main"><w:body/></w:document>'
            ),
        )
        if external_relationship_type is not None:
            package.writestr(
                "word/_rels/document.xml.rels",
                (
                    '<?xml version="1.0" encoding="UTF-8"?>'
                    "<Relationships "
                    'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                    '<Relationship Id="rIdExternal" '
                    f'Type="{external_relationship_type}" '
                    f'Target="{external_target}" TargetMode="External"/>'
                    "</Relationships>"
                ),
            )
        for part_name, part_payload in (additional_parts or {}).items():
            package.writestr(part_name, part_payload)
    return payload.getvalue()


def test_excel_inspection_reports_hidden_content_and_uncalculated_formulas() -> None:
    workbook = Workbook()
    visible = workbook.active
    visible.title = "BoQ"
    visible["A1"] = 10
    visible["A2"] = 20
    visible["A3"] = "=SUM(A1:A2)"
    visible.row_dimensions[2].hidden = True
    visible.column_dimensions["B"].hidden = True
    visible.column_dimensions.group("P", "AC", hidden=True)
    hidden = workbook.create_sheet("Hidden assumptions")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "manual coefficient"
    output = BytesIO()
    workbook.save(output)

    manifest = inspect_intake("boq.xlsx", output.getvalue(), Settings(app_env="test"))
    codes = {finding.code for finding in manifest.findings}
    assert "HIDDEN_EXCEL_SHEET" in codes
    assert "HIDDEN_EXCEL_DIMENSIONS" in codes
    assert "EXCEL_FORMULA_CACHE_MISSING" in codes
    assert not manifest.all_files_processed
    assert {sheet.name for sheet in manifest.entries[0].sheets} == {
        "BoQ",
        "Hidden assumptions",
    }
    boq = next(sheet for sheet in manifest.entries[0].sheets if sheet.name == "BoQ")
    assert boq.hidden_column_count == 15


def test_excel_inspection_blocks_broken_formula_and_cached_error() -> None:
    payload = _workbook_with_cached_formula_error(
        formula="=SUM(#REF!)",
        cached_error="#REF!",
    )
    manifest = inspect_intake("cost.xlsx", payload, Settings(app_env="test"))

    finding = next(
        finding for finding in manifest.findings if finding.code == "EXCEL_FORMULA_ERROR"
    )
    assert not manifest.all_files_processed
    assert finding.severity.value == "BLOCKER"
    assert finding.details == {
        "sheet": "Cost",
        "count": 1,
        "cells": [{"coordinate": "A2", "errors": ["#REF!"]}],
        "cells_truncated": False,
    }
    inspection = manifest.entries[0].sheets[0]
    assert inspection.formula_error_cell_count == 1
    assert inspection.formula_without_cached_value_count == 0
    assert not any(finding.code == "EXCEL_FORMULA_CACHE_MISSING" for finding in manifest.findings)


def test_excel_inspection_blocks_cached_formula_error_without_error_token() -> None:
    payload = _workbook_with_cached_formula_error(
        formula="=1/0",
        cached_error="#DIV/0!",
    )
    manifest = inspect_intake("cost.xlsx", payload, Settings(app_env="test"))

    finding = next(
        finding for finding in manifest.findings if finding.code == "EXCEL_FORMULA_ERROR"
    )
    assert finding.details["cells"] == [{"coordinate": "A2", "errors": ["#DIV/0!"]}]
    assert not manifest.all_files_processed


def test_excel_inspection_does_not_confuse_text_literal_with_formula_error() -> None:
    workbook = Workbook()
    workbook.active["A1"] = '=IF(B1="#REF!",1,0)'
    output = BytesIO()
    workbook.save(output)

    manifest = inspect_intake("cost.xlsx", output.getvalue(), Settings(app_env="test"))

    assert not any(finding.code == "EXCEL_FORMULA_ERROR" for finding in manifest.findings)


def test_excel_inspection_blocks_non_formula_error_value() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cost"
    sheet["A1"] = "#DIV/0!"
    sheet["A1"].data_type = "e"
    output = BytesIO()
    workbook.save(output)

    manifest = inspect_intake("cost.xlsx", output.getvalue(), Settings(app_env="test"))

    finding = next(finding for finding in manifest.findings if finding.code == "EXCEL_CELL_ERROR")
    assert not manifest.all_files_processed
    assert finding.details == {
        "sheet": "Cost",
        "count": 1,
        "cells": [{"coordinate": "A1", "error": "#DIV/0!"}],
        "cells_truncated": False,
    }
    assert manifest.entries[0].sheets[0].non_formula_error_cell_count == 1


def test_archive_path_traversal_is_blocked() -> None:
    payload = BytesIO()
    with ZipFile(payload, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("../outside.txt", "unsafe")
        archive.writestr("safe.txt", "safe")
    manifest = inspect_intake("package.zip", payload.getvalue(), Settings(app_env="test"))
    assert not manifest.all_files_processed
    assert any(finding.code == "ARCHIVE_PATH_TRAVERSAL" for finding in manifest.findings)
    assert any(entry.archive_path.endswith("safe.txt") for entry in manifest.entries)


def test_office_container_is_preflighted_before_excel_parser() -> None:
    payload = BytesIO()
    with ZipFile(payload, "w", compression=ZIP_DEFLATED) as package:
        package.writestr("xl/workbook.xml", "A" * 100_000)
    manifest = inspect_intake(
        "boq.xlsx",
        payload.getvalue(),
        Settings(app_env="test", max_archive_compression_ratio=2),
    )
    assert not manifest.all_files_processed
    assert any(finding.code == "OFFICE_COMPRESSION_RATIO_EXCEEDED" for finding in manifest.findings)


@pytest.mark.parametrize("filename", ["fake.docx", "fake.pptx"])
def test_office_structure_blocks_plain_zip_disguised_as_document(filename: str) -> None:
    payload = BytesIO()
    with ZipFile(payload, "w", compression=ZIP_DEFLATED) as package:
        package.writestr("not-an-office-document.txt", "safe text")

    manifest = inspect_intake(filename, payload.getvalue(), Settings(app_env="test"))

    assert not manifest.all_files_processed
    assert manifest.entries[0].corrupt
    assert any(finding.code == "CORRUPT_OFFICE_STRUCTURE" for finding in manifest.findings)


def test_office_structure_reports_external_hyperlink_without_blocking() -> None:
    payload = _minimal_docx(
        external_relationship_type=(
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink"
        ),
    )

    manifest = inspect_intake("terms.docx", payload, Settings(app_env="test"))

    assert manifest.all_files_processed
    finding = next(
        finding for finding in manifest.findings if finding.code == "OFFICE_EXTERNAL_HYPERLINK"
    )
    assert finding.severity.value == "WARNING"
    assert finding.details["count"] == 1
    assert manifest.entries[0].external_hyperlink_count == 1
    assert manifest.entries[0].external_dependency_count == 0


def test_office_structure_blocks_non_web_external_hyperlink() -> None:
    payload = _minimal_docx(
        external_relationship_type=(
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink"
        ),
        external_target="file:///C:/local/reference.docx",
    )

    manifest = inspect_intake("terms.docx", payload, Settings(app_env="test"))

    assert not manifest.all_files_processed
    assert any(finding.code == "OFFICE_EXTERNAL_DEPENDENCY" for finding in manifest.findings)
    assert manifest.entries[0].external_hyperlink_count == 0
    assert manifest.entries[0].external_dependency_count == 1


def test_office_structure_blocks_non_hyperlink_external_dependency() -> None:
    payload = _minimal_docx(
        external_relationship_type=(
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/attachedTemplate"
        ),
    )

    manifest = inspect_intake("terms.docx", payload, Settings(app_env="test"))

    assert not manifest.all_files_processed
    finding = next(
        finding for finding in manifest.findings if finding.code == "OFFICE_EXTERNAL_DEPENDENCY"
    )
    assert finding.severity.value == "BLOCKER"
    assert manifest.entries[0].external_dependency_count == 1


def test_office_structure_rejects_dtd_before_domain_use() -> None:
    payload = _minimal_docx(
        document_xml=(
            b'<?xml version="1.0"?>'
            b'<!DOCTYPE w:document [<!ENTITY unsafe "expanded">]>'
            b'<w:document xmlns:w="http://schemas.openxmlformats.org/'
            b'wordprocessingml/2006/main"><w:body>&unsafe;</w:body></w:document>'
        )
    )

    manifest = inspect_intake("terms.docx", payload, Settings(app_env="test"))

    assert not manifest.all_files_processed
    assert manifest.entries[0].corrupt
    assert any(finding.code == "CORRUPT_OFFICE_STRUCTURE" for finding in manifest.findings)


def test_office_structure_rejects_dtd_in_secondary_xml_part() -> None:
    payload = _minimal_docx(
        additional_parts={
            "customXml/item1.xml": (
                b'<?xml version="1.0"?>'
                b'<!DOCTYPE item [<!ENTITY unsafe "expanded">]>'
                b"<item>&unsafe;</item>"
            )
        }
    )

    manifest = inspect_intake("terms.docx", payload, Settings(app_env="test"))

    assert not manifest.all_files_processed
    assert manifest.entries[0].corrupt
    assert any(finding.code == "CORRUPT_OFFICE_STRUCTURE" for finding in manifest.findings)


def test_intake_manifest_rejects_green_status_with_blocker() -> None:
    valid = inspect_intake(
        "terms.docx",
        _minimal_docx(),
        Settings(app_env="test"),
    )
    payload = valid.model_dump(mode="json")
    payload["findings"] = [
        {
            "code": "TAMPERED_BLOCKER",
            "severity": "BLOCKER",
            "archive_path": "terms.docx",
            "message": "Contradictory persisted finding",
            "details": {},
        }
    ]

    with pytest.raises(ValueError, match="all_files_processed contradicts"):
        IntakeManifest.model_validate(payload)


def test_local_object_store_is_content_addressed_and_deduplicated(tmp_path: Path) -> None:
    store = LocalObjectStore(tmp_path / "objects")
    first = store.put(BytesIO(b"same content"))
    second = store.put(BytesIO(b"same content"))
    assert first == second
    with store.open(first.object_hash) as stream:
        assert stream.read() == b"same content"
    stored_files = [path for path in (tmp_path / "objects").rglob("*") if path.is_file()]
    assert len(stored_files) == 1
    stored_files[0].write_bytes(b"tampered content")
    with (
        pytest.raises(RuntimeError, match="fails SHA-256 verification"),
        store.open(first.object_hash),
    ):
        pass


def test_stream_intake_rejects_signature_mismatch_without_materializing_root() -> None:
    manifest = inspect_intake_stream(
        "disguised.pdf",
        BytesIO(b"this is not a PDF"),
        Settings(app_env="test"),
    )
    assert not manifest.all_files_processed
    assert any(finding.code == "FILE_SIGNATURE_MISMATCH" for finding in manifest.findings)


def test_object_store_size_limit_leaves_no_partial_object(tmp_path: Path) -> None:
    store = LocalObjectStore(tmp_path / "objects")
    with pytest.raises(ValueError, match="exceeds configured limit"):
        store.put(BytesIO(b"too large"), max_bytes=3)
    assert not [path for path in (tmp_path / "objects").rglob("*") if path.is_file()]
