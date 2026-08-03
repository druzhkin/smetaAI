# ruff: noqa: RUF001

from __future__ import annotations

import json
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from docx import Document
from openpyxl import load_workbook

from tenderguard import cli
from tenderguard.application.analysis_reporting import (
    BoqAnalysisRow,
    analysis_report_hash,
    build_analysis_from_extraction,
    build_analysis_from_price_matrix,
)
from tenderguard.application.pricing import (
    BoqPriceMatrixRowView,
    BoqPriceMatrixView,
    BoqPriceNameMatchView,
    BoqProposedPriceView,
    BoqSourcePriceView,
    NormalizedPriceView,
)
from tenderguard.domain.boq_spreadsheet import (
    BoqCellEvidence,
    BoqRowCandidate,
    BoqXlsxExtractionResult,
)
from tenderguard.domain.common import utc_now
from tenderguard.domain.enums import (
    MatchClass,
    PriceEvidenceClass,
    PriceSourceType,
)
from tenderguard.domain.models import PriceSourceReference
from tenderguard.infrastructure.boq_analysis_export import (
    build_boq_analysis_docx,
    build_boq_analysis_workbook,
    verify_boq_analysis_docx,
    verify_boq_analysis_workbook,
)


def _extraction() -> BoqXlsxExtractionResult:
    return BoqXlsxExtractionResult(
        status="BLOCKED",
        profile_version_id="profile-v1",
        profile_content_hash="a" * 64,
        root_object_sha256="b" * 64,
        workbook_object_sha256="c" * 64,
        archive_path="source.xlsx",
        worksheet_name="BoQ",
        candidates=(
            BoqRowCandidate(
                provisional_candidate_id="boq-candidate-" + "d" * 24,
                worksheet_name="BoQ",
                row_number=2,
                source_position_id="1",
                description="=HYPERLINK(\"https://invalid.example\",\"unsafe\")",
                unit="m",
                quantity=Decimal("10"),
                cells={
                    "description": BoqCellEvidence(
                        coordinate="B2",
                        value_kind="TEXT",
                        source_literal="=HYPERLINK(\"https://invalid.example\",\"unsafe\")",
                    )
                },
                blockers=("NOMENCLATURE_MATCH_MISSING",),
            ),
        ),
        global_blockers=("INTAKE_MANIFEST_BLOCKED",),
        extracted_at=utc_now(),
    )


def _source(
    *,
    group: PriceSourceType,
    evidence_class: PriceEvidenceClass,
    amount: str,
) -> BoqSourcePriceView:
    quote_id = f"quote-{group.value.lower()}"
    return BoqSourcePriceView(
        quote_id=quote_id,
        evidence_class=evidence_class,
        source_reference=PriceSourceReference(
            source_type=group,
            display_name=f"Source {group.value}",
            source_item_name=f"Source-side pipe {group.value}",
            source_record_id=f"record-{group.value}",
            source_uri=f"https://example.com/{group.value.lower()}",
        ),
        source_observation_id=f"observation-{group.value}",
        source_origin_id=f"origin-{group.value}",
        source_locator=f"https://example.com/{group.value.lower()}",
        source_document_revision_id=f"revision-{group.value}",
        observed_at=utc_now(),
        quote_date=date(2026, 8, 1),
        valid_until=date(2026, 8, 31),
        available=True,
        lead_time_days=5,
        raw_amount=Decimal(amount),
        raw_currency="RUB",
        raw_unit="m",
        normalized_prices=(
            NormalizedPriceView(
                normalized_price_id=f"normalized-{group.value}",
                quote_id=quote_id,
                amount_per_unit=Decimal(amount),
                currency="RUB",
                unit="m",
                formula_hash="f" * 64,
                policy_version_id="price-policy-v1",
            ),
        ),
        technical_attributes={"diameter": "100"},
    )


def _matrix() -> BoqPriceMatrixView:
    won = _source(
        group=PriceSourceType.WON_TENDER,
        evidence_class=PriceEvidenceClass.INTERNAL_HISTORY,
        amount="100",
    )
    fgis = _source(
        group=PriceSourceType.FGIS_CS,
        evidence_class=PriceEvidenceClass.OFFICIAL_OR_PRIMARY,
        amount="101",
    )
    market = _source(
        group=PriceSourceType.SUPPLIER_WEBSITE,
        evidence_class=PriceEvidenceClass.INDEPENDENT_MARKET,
        amount="102",
    )
    return BoqPriceMatrixView(
        project_id="project-1",
        generated_at=utc_now(),
        rows=(
            BoqPriceMatrixRowView(
                row_id="line-1:1",
                boq_line_id="line-1",
                line_key="1",
                wbs_node_id="wbs-1",
                work_code="PIPE",
                boq_item_name="Pipe material",
                boq_unit="m",
                quantity=Decimal("2"),
                quantity_status="VERIFIED",
                item_id="pipe-source",
                cost_category="MATERIAL",
                basis_kind="MARKET",
                row_status="VERIFIED",
                blockers=(),
                name_match=BoqPriceNameMatchView(
                    match_id="match-1",
                    status="VERIFIED",
                    match_class=MatchClass.EXACT,
                    boq_item_name="Pipe material",
                    source_item_id="pipe-source",
                    canonical_item_id="pipe-canonical",
                    source_attributes={"diameter": "100"},
                    canonical_attributes={"diameter": "100"},
                    mismatched_attributes=(),
                    missing_attributes=(),
                    catalog_version_id="catalog-v1",
                    assessment_method="EXACT_CRITICAL_ATTRIBUTES",
                ),
                won_tender_prices=(won,),
                fgis_cs_prices=(fgis,),
                market_prices=(market,),
                other_prices=(),
                proposed_price=BoqProposedPriceView(
                    status="VERIFIED",
                    workflow_status="VERIFIED",
                    amount_per_unit=Decimal("101"),
                    currency="RUB",
                    unit="m",
                    decision_id="decision-1",
                    as_of=date(2026, 8, 1),
                    selection_method="approved-median",
                    normalized_price_ids=(
                        "normalized-WON_TENDER",
                        "normalized-FGIS_CS",
                        "normalized-SUPPLIER_WEBSITE",
                    ),
                    rationale=("Approved deterministic selection.",),
                ),
            ),
        ),
        blocked_row_count=0,
        release_warning="The bid release gate remains authoritative.",
    )


def test_extraction_report_withholds_prices_and_escapes_excel_formula_text() -> None:
    report = build_analysis_from_extraction(
        extraction=_extraction(),
        project_id="project-1",
        project_code="ALABUGA-4527946",
    )

    assert report.analysis_status == "BLOCKED"
    assert report.final_total is None
    assert report.rows[0].proposed_amount_per_unit is None

    workbook_content = build_boq_analysis_workbook(report)
    assert build_boq_analysis_workbook(report) == workbook_content
    verify_boq_analysis_workbook(workbook_content, report)
    workbook = load_workbook(BytesIO(workbook_content), data_only=False)
    try:
        assert workbook.sheetnames == [
            "ВОР_с_оценкой",
            "Источники",
            "Сопоставления",
            "Блокировки",
            "Метаданные",
        ]
        main = workbook["ВОР_с_оценкой"]
        headers = {cell.value: cell.column for cell in main[4]}
        assert main.cell(5, headers["Цена системы"]).value is None
        assert main.cell(5, headers["Стоимость строки"]).value is None
        assert main.cell(5, headers["Наименование ВОР/ТЗ"]).data_type != "f"
        assert main.cell(5, headers["Наименование ВОР/ТЗ"]).value.startswith("'=")
    finally:
        workbook.close()

    docx_content = build_boq_analysis_docx(report)
    assert build_boq_analysis_docx(report) == docx_content
    verify_boq_analysis_docx(docx_content, report)
    document = Document(BytesIO(docx_content))
    text = "\n".join(paragraph.text for paragraph in document.paragraphs)
    assert report.project_code in text
    assert analysis_report_hash(report) in text


def test_price_matrix_report_retains_source_names_and_refuses_offline_release() -> None:
    blocked = build_analysis_from_price_matrix(
        matrix=_matrix(),
        project_code="ALABUGA-4527946",
        release_state="EXPERT_REVIEW",
    )

    assert blocked.analysis_status == "BLOCKED"
    assert blocked.final_total is None
    assert blocked.rows[0].line_amount == Decimal("202")
    assert {source.source_group for source in blocked.rows[0].sources} == {
        "WON_TENDER",
        "FGIS_CS",
        "MARKET",
    }
    assert {source.source_item_name for source in blocked.rows[0].sources} == {
        "Source-side pipe WON_TENDER",
        "Source-side pipe FGIS_CS",
        "Source-side pipe SUPPLIER_WEBSITE",
    }

    forged_release = build_analysis_from_price_matrix(
        matrix=_matrix(),
        project_code="ALABUGA-4527946",
        release_state="APPROVED_FOR_BID",
        calculation_snapshot_id="snapshot-1",
        final_total=Decimal("202"),
        final_currency="RUB",
    )
    assert forged_release.analysis_status == "BLOCKED"
    assert forged_release.final_total is None
    assert "GOVERNED_RELEASE_EXPORT_NOT_IMPLEMENTED" in forged_release.global_blockers


def test_empty_extraction_creates_no_phantom_spreadsheet_rows() -> None:
    extraction = _extraction().model_copy(update={"candidates": ()})
    report = build_analysis_from_extraction(
        extraction=extraction,
        project_id="project-1",
        project_code="ALABUGA-EMPTY",
    )

    workbook_content = build_boq_analysis_workbook(report)
    workbook = load_workbook(BytesIO(workbook_content), data_only=False)
    try:
        assert workbook["ВОР_с_оценкой"].max_row == 4
        assert workbook["Источники"].max_row == 3
        assert workbook["Сопоставления"].max_row == 3
    finally:
        workbook.close()


def test_blocked_analysis_row_rejects_proposed_amount() -> None:
    with pytest.raises(ValueError, match="blocked analysis row"):
        BoqAnalysisRow(
            row_id="row-1",
            boq_line_id="line-1",
            line_key="1",
            wbs_node_id="wbs-1",
            work_code="PIPE",
            boq_item_name="Pipe",
            boq_unit="m",
            quantity=Decimal("1"),
            quantity_status="UNVERIFIED",
            item_id="pipe",
            row_status="BLOCKED",
            blockers=("PRICE_DECISION_MISSING",),
            proposed_amount_per_unit=Decimal("100"),
            proposed_currency="RUB",
            proposed_unit="m",
            price_decision_id="decision-1",
            price_as_of=date(2026, 8, 1),
            selection_method="unsafe",
            rationale=("unsafe",),
        )


def test_cli_exports_exclusive_analysis_package(tmp_path, capsys) -> None:
    input_path = tmp_path / "probe.json"
    output_dir = tmp_path / "analysis"
    input_path.write_text(_extraction().model_dump_json(indent=2), encoding="utf-8")

    assert (
        cli.export_boq_analysis(
            input_path=input_path,
            input_kind="xlsx-extraction",
            project_id="project-1",
            project_code="ALABUGA-4527946",
            release_state="BLOCKED",
            output_dir=output_dir,
        )
        == 0
    )
    manifest = json.loads(capsys.readouterr().out)
    assert manifest["analysis_status"] == "BLOCKED"
    assert len(manifest["artifacts"]) == 2
    assert all((output_dir / item["filename"]).is_file() for item in manifest["artifacts"])
    assert (output_dir / "manifest.json").is_file()

    assert (
        cli.export_boq_analysis(
            input_path=input_path,
            input_kind="xlsx-extraction",
            project_id="project-1",
            project_code="ALABUGA-4527946",
            release_state="BLOCKED",
            output_dir=output_dir,
        )
        == 2
    )
    assert json.loads(capsys.readouterr().out)["code"] == "ANALYSIS_OUTPUT_ALREADY_EXISTS"


def test_cli_removes_partial_package_when_manifest_write_fails(
    tmp_path,
    capsys,
    monkeypatch,
) -> None:
    input_path = tmp_path / "probe.json"
    output_dir = tmp_path / "analysis"
    input_path.write_text(_extraction().model_dump_json(indent=2), encoding="utf-8")

    def fail_manifest_write(destination, payload) -> None:
        del destination, payload
        raise OSError("simulated manifest storage failure")

    monkeypatch.setattr(cli, "_write_text_exclusive", fail_manifest_write)

    assert (
        cli.export_boq_analysis(
            input_path=input_path,
            input_kind="xlsx-extraction",
            project_id="project-1",
            project_code="ALABUGA-4527946",
            release_state="BLOCKED",
            output_dir=output_dir,
        )
        == 2
    )
    assert json.loads(capsys.readouterr().out)["artifacts_created"] is False
    assert not output_dir.exists()
    assert not tuple(tmp_path.glob(".analysis.staging-*"))
